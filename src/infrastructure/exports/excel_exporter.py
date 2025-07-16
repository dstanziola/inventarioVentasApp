"""
ExcelExporter - Exportador especializado para formato Excel.

Este módulo maneja la creación de archivos Excel (.xlsx) con formato profesional,
incluyendo estilos corporativos, formateo de datos y plantillas personalizadas.

FUNCIONALIDADES:
- Creación de workbooks Excel con múltiples hojas
- Aplicación de estilos y formatos profesionales
- Inserción de logos e información corporativa
- Formateo automático de columnas y datos
- Generación de gráficos y resúmenes

DEPENDENCIAS:
- openpyxl: Librería principal para manipulación Excel
- xlsxwriter: Funcionalidades avanzadas de formato (opcional)

ARQUITECTURA:
- Infrastructure Layer: Implementación concreta de exportación
- Single Responsibility: Solo maneja exportación Excel
- Open/Closed: Extensible para nuevos tipos de reportes Excel

SPRINT 2: Implementación TDD con tests previos
Autor: Sistema de Inventario Copy Point S.A.
Versión: 1.0.0 - Sprint 2
Fecha: 2025-07-12
"""

import os
import logging
from datetime import datetime, date
from decimal import Decimal
from typing import Dict, Any, List, Optional, Union

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.chart import PieChart, BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None

# Configurar logging
logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Exportador especializado para archivos Excel con formato profesional.
    
    RESPONSABILIDADES:
    - Crear workbooks Excel con formato corporativo
    - Aplicar estilos y diseño profesional
    - Insertar datos con formato apropiado
    - Generar gráficos y resúmenes automáticos
    - Optimizar columnas y presentación
    
    PATRONES:
    - Template Method: Proceso estándar de creación Excel
    - Strategy Pattern: Diferentes estilos según tipo de reporte
    - Builder Pattern: Construcción progresiva de workbook
    """
    
    def __init__(self):
        """
        Inicializar exportador Excel.
        
        Raises:
            ImportError: Si openpyxl no está disponible
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl es requerido para ExcelExporter. "
                "Instalar con: pip install openpyxl>=3.0.0"
            )
        
        # Configuración de estilos corporativos
        self.corporate_colors = {
            'primary': '1F4E79',      # Azul corporativo
            'secondary': '70AD47',    # Verde corporativo  
            'accent': 'C55A11',       # Naranja corporativo
            'light_gray': 'F2F2F2',   # Gris claro
            'dark_gray': '595959'     # Gris oscuro
        }
        
        # Configuración de fuentes
        self.fonts = {
            'title': Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
            'header': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'data': Font(name='Calibri', size=10),
            'summary': Font(name='Calibri', size=11, bold=True)
        }
        
        # Información corporativa
        self.company_info = {
            'nombre': 'Copy Point S.A.',
            'ruc': '888-888-8888',
            'direccion': 'Las Lajas, Las Cumbres, Panamá',
            'telefono': '6666-6666',
            'email': 'copy.point@gmail.com'
        }
        
        logger.info("ExcelExporter inicializado con estilos corporativos")
    
    def create_movements_workbook(self, data: Dict[str, Any], file_path: str) -> None:
        """
        Crear workbook Excel para movimientos de inventario.
        
        Args:
            data: Datos formateados con plantilla aplicada
            file_path: Ruta donde guardar el archivo Excel
        
        Raises:
            ValueError: Si los datos no tienen formato correcto
            IOError: Si no se puede escribir el archivo
            Exception: Para otros errores de creación
        """
        try:
            # Validar datos de entrada
            self._validate_workbook_data(data)
            
            # Crear workbook
            workbook = Workbook()
            
            # Eliminar hoja por defecto
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            
            # Crear hojas según contenido
            self._create_data_sheet(workbook, data)
            self._create_summary_sheet(workbook, data)
            
            # Aplicar protección y configuraciones finales
            self._apply_workbook_settings(workbook)
            
            # Guardar archivo
            workbook.save(file_path)
            logger.info(f"Workbook Excel creado exitosamente: {file_path}")
            
        except ValueError as e:
            logger.error(f"Datos inválidos para workbook: {e}")
            raise
        except IOError as e:
            logger.error(f"Error escribiendo archivo Excel: {e}")
            raise IOError(f"No se pudo guardar archivo Excel: {e}")
        except Exception as e:
            logger.error(f"Error inesperado creando workbook: {e}")
            raise Exception(f"Error creando workbook Excel: {e}")
    
    def _validate_workbook_data(self, data: Dict[str, Any]) -> None:
        """
        Validar que los datos tienen la estructura correcta para Excel.
        
        Args:
            data: Datos a validar
        
        Raises:
            ValueError: Si la estructura no es válida
        """
        if not isinstance(data, dict):
            raise ValueError("Data debe ser un diccionario")
        
        required_keys = ['title', 'data', 'filters']
        for key in required_keys:
            if key not in data:
                raise ValueError(f"Clave requerida '{key}' faltante en data")
        
        if not isinstance(data['data'], list):
            raise ValueError("data['data'] debe ser una lista")
        
        logger.debug(f"Datos validados: {len(data['data'])} registros")
    
    def _create_data_sheet(self, workbook: Workbook, data: Dict[str, Any]) -> None:
        """
        Crear hoja principal con datos de movimientos.
        
        Args:
            workbook: Workbook de Excel
            data: Datos formateados
        """
        # Crear hoja de datos
        sheet = workbook.create_sheet("Movimientos", 0)
        
        # Agregar header corporativo
        self._add_corporate_header(sheet, data)
        
        # Agregar datos de movimientos
        if data['data']:
            self._add_movements_table(sheet, data['data'])
        else:
            self._add_empty_data_message(sheet)
        
        # Aplicar formato y ajustes
        self._format_data_sheet(sheet)
        
        logger.debug("Hoja de datos creada con formato corporativo")
    
    def _create_summary_sheet(self, workbook: Workbook, data: Dict[str, Any]) -> None:
        """
        Crear hoja de resumen con estadísticas.
        
        Args:
            workbook: Workbook de Excel
            data: Datos con resumen incluido
        """
        sheet = workbook.create_sheet("Resumen", 1)
        
        # Header corporativo para resumen
        self._add_corporate_header(sheet, data, sheet_type="resumen")
        
        # Agregar estadísticas
        summary_data = data.get('summary', {})
        self._add_summary_statistics(sheet, summary_data)
        
        # Agregar gráficos si hay datos
        if data['data']:
            self._add_summary_charts(sheet, data['data'])
        
        # Formato final
        self._format_summary_sheet(sheet)
        
        logger.debug("Hoja de resumen creada con gráficos")
    
    def _add_corporate_header(self, sheet, data: Dict[str, Any], sheet_type: str = "data") -> None:
        """
        Agregar header corporativo a la hoja.
        
        Args:
            sheet: Hoja de Excel
            data: Datos con título e información
            sheet_type: Tipo de hoja para personalizar header
        """
        # Título principal
        sheet['A1'] = self.company_info['nombre']
        sheet['A1'].font = Font(name='Calibri', size=18, bold=True, color=self.corporate_colors['primary'])
        sheet.merge_cells('A1:G1')
        
        # Información corporativa
        sheet['A2'] = f"RUC: {self.company_info['ruc']} | Tel: {self.company_info['telefono']}"
        sheet['A2'].font = self.fonts['data']
        sheet.merge_cells('A2:G2')
        
        # Título del reporte
        title = data.get('title', 'Reporte de Inventario')
        sheet['A4'] = title
        sheet['A4'].font = self.fonts['title']
        sheet['A4'].fill = PatternFill(start_color=self.corporate_colors['primary'], 
                                      end_color=self.corporate_colors['primary'], 
                                      fill_type='solid')
        sheet.merge_cells('A4:G4')
        
        # Información de filtros aplicados
        filters_info = self._format_filters_info(data.get('filters', {}))
        sheet['A5'] = f"Filtros aplicados: {filters_info}"
        sheet['A5'].font = self.fonts['data']
        sheet.merge_cells('A5:G5')
        
        # Fecha de generación
        sheet['A6'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        sheet['A6'].font = self.fonts['data']
        sheet.merge_cells('A6:G6')
    
    def _add_movements_table(self, sheet, movements_data: List[Dict[str, Any]]) -> None:
        """
        Agregar tabla de movimientos con formato profesional.
        
        Args:
            sheet: Hoja de Excel
            movements_data: Lista de movimientos formateados
        """
        if not movements_data:
            return
        
        # Determinar posición inicial (después del header)
        start_row = 8
        
        # Headers de columnas
        headers = list(movements_data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row, column=col, value=header)
            cell.font = self.fonts['header']
            cell.fill = PatternFill(start_color=self.corporate_colors['primary'],
                                   end_color=self.corporate_colors['primary'],
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos de movimientos
        for row_idx, movement in enumerate(movements_data, start_row + 1):
            for col_idx, value in enumerate(movement.values(), 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.fonts['data']
                
                # Aplicar formato específico según columna
                header = headers[col_idx - 1]
                if 'cantidad' in header.lower() or 'stock' in header.lower():
                    cell.alignment = Alignment(horizontal='right')
                elif 'fecha' in header.lower():
                    cell.alignment = Alignment(horizontal='center')
        
        # Crear tabla Excel con estilo
        table_range = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(movements_data)}"
        table = Table(displayName="MovimientosTable", ref=table_range)
        
        # Estilo de tabla
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        sheet.add_table(table)
        logger.debug(f"Tabla de movimientos agregada: {len(movements_data)} registros")
    
    def _add_empty_data_message(self, sheet) -> None:
        """Agregar mensaje cuando no hay datos para mostrar."""
        sheet['A8'] = "No hay datos para mostrar con los filtros aplicados"
        sheet['A8'].font = Font(name='Calibri', size=12, italic=True, color='999999')
        sheet.merge_cells('A8:G8')
        sheet['A8'].alignment = Alignment(horizontal='center', vertical='center')
    
    def _add_summary_statistics(self, sheet, summary: Dict[str, Any]) -> None:
        """
        Agregar estadísticas de resumen a la hoja.
        
        Args:
            sheet: Hoja de Excel
            summary: Datos de resumen
        """
        start_row = 8
        
        # Título de resumen
        sheet.cell(row=start_row, column=1, value="RESUMEN EJECUTIVO").font = self.fonts['summary']
        sheet.merge_cells(f'A{start_row}:D{start_row}')
        
        # Estadísticas principales
        stats = [
            ("Total de Movimientos:", summary.get('total_movimientos', 0)),
            ("Total Entradas:", summary.get('total_entradas', 0)),
            ("Total Ajustes:", summary.get('total_ajustes', 0)),
            ("Valor Total:", summary.get('valor_total', 'B/. 0.00'))
        ]
        
        for idx, (label, value) in enumerate(stats, start_row + 2):
            sheet.cell(row=idx, column=1, value=label).font = self.fonts['data']
            cell_value = sheet.cell(row=idx, column=2, value=value)
            cell_value.font = self.fonts['summary']
            
            if isinstance(value, (int, float)) and value > 0:
                cell_value.fill = PatternFill(start_color=self.corporate_colors['secondary'],
                                            end_color=self.corporate_colors['secondary'],
                                            fill_type='solid')
    
    def _add_summary_charts(self, sheet, movements_data: List[Dict[str, Any]]) -> None:
        """
        Agregar gráficos de resumen.
        
        Args:
            sheet: Hoja de Excel
            movements_data: Datos para generar gráficos
        """
        try:
            # Contar movimientos por tipo
            type_counts = {}
            for movement in movements_data:
                tipo = movement.get('Tipo', 'OTRO')
                type_counts[tipo] = type_counts.get(tipo, 0) + 1
            
            if type_counts:
                # Crear gráfico de torta para tipos de movimiento
                chart = PieChart()
                chart.title = "Distribución por Tipo de Movimiento"
                
                # Datos para el gráfico (simplificado)
                start_row = 15
                sheet.cell(row=start_row, column=5, value="Tipo")
                sheet.cell(row=start_row, column=6, value="Cantidad")
                
                for idx, (tipo, count) in enumerate(type_counts.items(), start_row + 1):
                    sheet.cell(row=idx, column=5, value=tipo)
                    sheet.cell(row=idx, column=6, value=count)
                
                # Configurar referencias del gráfico
                data = Reference(sheet, min_col=6, min_row=start_row, 
                               max_row=start_row + len(type_counts))
                labels = Reference(sheet, min_col=5, min_row=start_row + 1, 
                                 max_row=start_row + len(type_counts))
                
                chart.add_data(data, titles_from_data=False)
                chart.set_categories(labels)
                
                # Agregar gráfico a la hoja
                sheet.add_chart(chart, "H8")
                
                logger.debug("Gráfico de resumen agregado exitosamente")
                
        except Exception as e:
            logger.warning(f"No se pudo crear gráfico de resumen: {e}")
    
    def _format_data_sheet(self, sheet) -> None:
        """Aplicar formato final a la hoja de datos."""
        # Auto-ajustar columnas
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar paneles (primera fila después del header)
        sheet.freeze_panes = 'A9'
        
        logger.debug("Formato aplicado a hoja de datos")
    
    def _format_summary_sheet(self, sheet) -> None:
        """Aplicar formato final a la hoja de resumen."""
        # Ajustar columnas del resumen
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        
        # Configurar vista
        sheet.sheet_view.showGridLines = False
        
        logger.debug("Formato aplicado a hoja de resumen")
    
    def _format_filters_info(self, filters: Dict[str, Any]) -> str:
        """
        Formatear información de filtros para mostrar en el reporte.
        
        Args:
            filters: Diccionario de filtros aplicados
        
        Returns:
            str: Texto formateado de filtros
        """
        if not filters:
            return "Ninguno"
        
        filter_parts = []
        
        for key, value in filters.items():
            if key in ['fecha_inicio', 'fecha_fin']:
                continue  # Manejado por separado
            
            if value and str(value).upper() not in ['TODOS', 'ALL', 'NINGUNO']:
                filter_parts.append(f"{key.replace('_', ' ').title()}: {value}")
        
        # Agregar rango de fechas si existe
        if 'fecha_inicio' in filters and 'fecha_fin' in filters:
            inicio = filters['fecha_inicio']
            fin = filters['fecha_fin']
            if inicio and fin:
                filter_parts.append(f"Período: {inicio} - {fin}")
        
        return "; ".join(filter_parts) if filter_parts else "Ninguno"
    
    def _apply_workbook_settings(self, workbook: Workbook) -> None:
        """
        Aplicar configuraciones finales al workbook.
        
        Args:
            workbook: Workbook de Excel
        """
        # Configuraciones generales del workbook
        workbook.properties.title = "Reporte de Inventario"
        workbook.properties.subject = "Sistema de Gestión de Inventario"
        workbook.properties.creator = self.company_info['nombre']
        workbook.properties.created = datetime.now()
        
        # Establecer hoja activa
        if len(workbook.worksheets) > 0:
            workbook.active = workbook.worksheets[0]
        
        logger.debug("Configuraciones finales aplicadas al workbook")
    
    def get_supported_formats(self) -> List[str]:
        """
        Obtener lista de formatos soportados por este exportador.
        
        Returns:
            List[str]: Formatos soportados
        """
        return ['.xlsx', '.xlsm']
    
    def validate_file_path(self, file_path: str) -> bool:
        """
        Validar que la ruta de archivo es válida para Excel.
        
        Args:
            file_path: Ruta del archivo
        
        Returns:
            bool: True si es válida
        """
        if not file_path:
            return False
        
        # Verificar extensión
        valid_extensions = self.get_supported_formats()
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext not in valid_extensions:
            return False
        
        # Verificar que el directorio padre existe o se puede crear
        directory = os.path.dirname(file_path)
        if directory and not os.path.exists(directory):
            try:
                os.makedirs(directory, exist_ok=True)
            except OSError:
                return False
        
        return True
